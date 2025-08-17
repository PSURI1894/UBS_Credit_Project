import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Create workbook
wb = Workbook()

# ==============================
# Sheet 1: Cluster Mapping
# ==============================
ws1 = wb.active
ws1.title = "Cluster Mapping"

headers = [
    "Cluster", "Avg Income", "Avg Credit Score", "Avg DTI", "Avg Spend",
    "Avg Late Payments", "Default Rate", "Churn Rate", "Suggested Segment"
]
ws1.append(headers)

# Example data (replace later with your own cluster_profile.xlsx)
example_data = [
    [0, 2100000, 780, 0.25, 65000, 0.5, 0.02, 0.05, "Premier"],
    [1, 1000000, 710, 0.40, 38000, 1.5, 0.08, 0.12, "Gold"],
    [2, 650000, 660, 0.55, 22000, 2.3, 0.15, 0.20, "Silver"],
    [3, 400000, 590, 0.70, 15000, 4.8, 0.30, 0.28, "High-Risk"],
]
for row in example_data:
    ws1.append(row)

# Style header row
for col in range(1, len(headers)+1):
    ws1.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
    ws1.cell(row=1, column=col).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

# ==============================
# Sheet 2: Segment Recommendations
# ==============================
ws2 = wb.create_sheet("Segment Recommendations")

ws2.append(["Segment", "Recommended UBS Actions"])

recommendations = {
    "Premier": [
        "Invite to Wealth Management services",
        "Offer higher credit limits",
        "Dedicated relationship manager outreach"
    ],
    "Gold": [
        "Controlled credit limit growth (based on repayment behavior)",
        "Auto-enroll in bill reminders",
        "Targeted loyalty rewards"
    ],
    "Silver": [
        "Financial wellness nudges (SMS/email alerts)",
        "Secured loans / credit builder products",
        "Encourage responsible credit usage"
    ],
    "High-Risk": [
        "Enhanced verification for new loans",
        "Restrict initial credit limits",
        "Offer payment-plan restructuring"
    ]
}

for seg, acts in recommendations.items():
    ws2.append([seg, " | ".join(acts)])

# Style sheet 2 header
ws2.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
ws2.cell(row=1, column=1).fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
ws2.cell(row=1, column=2).font = Font(bold=True, color="FFFFFF")
ws2.cell(row=1, column=2).fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")

# Save workbook
wb.save("ubs_cluster_mapping_with_recommendations.xlsx")
print("File saved as ubs_cluster_mapping_with_recommendations.xlsx")
